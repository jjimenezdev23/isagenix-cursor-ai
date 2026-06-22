from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_FILE = "toastie_costing_system.xlsx"


INGREDIENTS = [
    ("Baguette", "pc", 1, 85, "1 baguette is used as 7 toastie servings."),
    ("Butter", "g", 200, 49, "Purchase pack is 200g."),
    ("Garlic", "bulb", 7, 35, "Purchase pack has 7 bulbs."),
    ("Cheese", "g", 185, 90, "Default recipe assumes the 185g pack covers 10 servings."),
    ("Parmesan Cheese", "g", 350, 212, "1/4 cup parmesan is estimated as 25g for 10 servings."),
    ("Parsley", "g", 100, 40, "Fresh chopped parsley estimate: 1 tbsp = 3.8g."),
    ("Brown Sugar", "g", 250, 20, "250g sugar covers 25 Sugar Rush servings."),
    ("Takeout Box", "pc", 10, 70, "Optional packaging: 1 box per serving."),
    ("Dine-in Box", "pc", 25, 60, "Optional packaging: 1 box per serving."),
]


RECIPE_USAGE_HEADERS = [
    "Product",
    "Package Option",
    "Selling Price",
    "Baguette (pc)",
    "Butter (g)",
    "Garlic (bulb)",
    "Cheese (g)",
    "Parmesan Cheese (g)",
    "Parsley (g)",
    "Brown Sugar (g)",
    "Takeout Box (pc)",
    "Dine-in Box (pc)",
]


BASE_USAGE = {
    "Garlic Cheese Toasties": {
        "Baguette (pc)": 1 / 7,
        "Butter (g)": 113 / 10,
        "Garlic (bulb)": 4 / 10,
        "Cheese (g)": 185 / 10,
        "Parmesan Cheese (g)": 25 / 10,
        "Parsley (g)": 11.4 / 10,
        "Brown Sugar (g)": 0,
    },
    "Sugar Rush Toasties": {
        "Baguette (pc)": 1 / 7,
        "Butter (g)": 10,
        "Garlic (bulb)": 0,
        "Cheese (g)": 0,
        "Parmesan Cheese (g)": 0,
        "Parsley (g)": 0,
        "Brown Sugar (g)": 250 / 25,
    },
}


PACKAGE_OPTIONS = [
    ("No Box", 0, 0),
    ("Dine-in", 0, 1),
    ("Takeout", 1, 0),
]


MARGINS = [0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50]


thin_gray = Side(style="thin", color="D9E2F3")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
header_fill = PatternFill("solid", fgColor="1F4E78")
subheader_fill = PatternFill("solid", fgColor="5B9BD5")
input_fill = PatternFill("solid", fgColor="FFF2CC")
formula_fill = PatternFill("solid", fgColor="E2F0D9")
note_fill = PatternFill("solid", fgColor="FCE4D6")
white_font = Font(color="FFFFFF", bold=True)
title_font = Font(size=16, bold=True, color="1F4E78")
section_font = Font(size=12, bold=True, color="1F4E78")


def style_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_table(ws, name, start_row, end_row, end_col):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header_row(ws, row_number=1):
    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def product_variants():
    variants = []
    for product in BASE_USAGE:
        for package_option, takeout_qty, dine_in_qty in PACKAGE_OPTIONS:
            usage = dict(BASE_USAGE[product])
            usage["Takeout Box (pc)"] = takeout_qty
            usage["Dine-in Box (pc)"] = dine_in_qty
            variants.append(
                {
                    "product": product,
                    "package_option": package_option,
                    "selling_price": "",
                    "usage": usage,
                }
            )
    return variants


def build_workbook():
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_intro = wb.create_sheet("Instructions")
    ws_ingredients = wb.create_sheet("Ingredients")
    ws_usage = wb.create_sheet("Recipe Usage")
    ws_breakdown = wb.create_sheet("Cost Breakdown")
    ws_summary = wb.create_sheet("Product Summary")
    ws_margins = wb.create_sheet("Margin Pricing")

    variants = product_variants()

    build_instructions(ws_intro)
    build_ingredients(ws_ingredients)
    build_recipe_usage(ws_usage, variants)
    build_cost_breakdown(ws_breakdown, variants)
    build_summary(ws_summary, variants)
    build_margin_pricing(ws_margins, variants)

    wb.active = wb.sheetnames.index("Instructions")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT_FILE)

    # Validate that the workbook can be opened after writing.
    load_workbook(OUTPUT_FILE, data_only=False)


def build_instructions(ws):
    ws["A1"] = "Toastie Costing System"
    ws["A1"].font = title_font

    ws["A3"] = "How to use"
    ws["A3"].font = section_font
    instructions = [
        "1. Update purchase quantities and prices in the Ingredients sheet.",
        "2. Update per-serving ingredient quantities in the Recipe Usage sheet if your recipe changes.",
        "3. Enter your chosen selling price in the Recipe Usage sheet.",
        "4. Review Product Summary for capital per serving, packaging cost, profit, and actual margin.",
        "5. Review Margin Pricing for exact and rounded selling price options from 20% to 50% margin.",
        "6. Yellow cells are editable inputs. Green cells are formulas that recalculate automatically in Excel or Google Sheets.",
    ]
    for row_idx, text in enumerate(instructions, 4):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A11"] = "Current assumptions"
    ws["A11"].font = section_font
    assumptions = [
        "1 baguette = 7 servings.",
        "Garlic Cheese Toasties use baguette, butter, garlic, cheese, parmesan, and fresh parsley.",
        "Sugar Rush Toasties use baguette, butter, and brown sugar.",
        "Shredded cheese default: 185g cheese pack divided by 10 servings = 18.5g per serving.",
        "Parmesan default: 25g divided by 10 servings = 2.5g per serving.",
        "Fresh chopped parsley estimate: 3 tbsp = 11.4g for 10 servings = 1.14g per serving.",
        "Boxes are optional and represented as No Box, Dine-in, and Takeout variants.",
        "All prices and costs should use the same currency.",
    ]
    for row_idx, text in enumerate(assumptions, 12):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A22"] = "Formula color guide"
    ws["A22"].font = section_font
    legend = [
        ("Yellow", "Input cells: enter prices, purchase quantities, recipe quantities, and chosen selling price."),
        ("Green", "Formula cells: calculated automatically."),
        ("Blue", "Table headers."),
        ("Orange", "Notes and assumptions."),
    ]
    for row_idx, (label, text) in enumerate(legend, 23):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=text)
    ws["A23"].fill = input_fill
    ws["A24"].fill = formula_fill
    ws["A25"].fill = header_fill
    ws["A25"].font = white_font
    ws["A26"].fill = note_fill

    set_column_widths(ws, [28, 90])
    style_range(ws, 1, 26, 1, 2)
    ws.sheet_view.showGridLines = False


def build_ingredients(ws):
    headers = [
        "Ingredient",
        "Measurement Unit",
        "Purchase Qty",
        "Purchase Cost",
        "Cost Per Unit",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws)

    for row_idx, (ingredient, unit, qty, cost, note) in enumerate(INGREDIENTS, 2):
        ws.cell(row=row_idx, column=1, value=ingredient)
        ws.cell(row=row_idx, column=2, value=unit)
        ws.cell(row=row_idx, column=3, value=qty)
        ws.cell(row=row_idx, column=4, value=cost)
        ws.cell(
            row=row_idx,
            column=5,
            value=f'=IF(OR(C{row_idx}="",D{row_idx}=""),"",D{row_idx}/C{row_idx})',
        )
        ws.cell(row=row_idx, column=6, value=note)
        ws.cell(row=row_idx, column=3).fill = input_fill
        ws.cell(row=row_idx, column=4).fill = input_fill
        ws.cell(row=row_idx, column=5).fill = formula_fill
        ws.cell(row=row_idx, column=6).fill = note_fill

    last_row = len(INGREDIENTS) + 1
    add_table(ws, "IngredientsTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [22, 18, 16, 16, 16, 64])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.0000"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_recipe_usage(ws, variants):
    ws.append(RECIPE_USAGE_HEADERS)
    style_header_row(ws)

    for variant in variants:
        row = [variant["product"], variant["package_option"], variant["selling_price"]]
        for header in RECIPE_USAGE_HEADERS[3:]:
            row.append(variant["usage"].get(header, 0))
        ws.append(row)

    last_row = len(variants) + 1
    add_table(ws, "RecipeUsageTable", 1, last_row, len(RECIPE_USAGE_HEADERS))
    style_range(ws, 1, last_row, 1, len(RECIPE_USAGE_HEADERS))
    set_column_widths(ws, [28, 16, 14, 14, 12, 14, 12, 20, 12, 16, 18, 18])
    for row in range(2, last_row + 1):
        for col in range(3, len(RECIPE_USAGE_HEADERS) + 1):
            ws.cell(row=row, column=col).fill = input_fill
            ws.cell(row=row, column=col).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_cost_breakdown(ws, variants):
    headers = [
        "Product",
        "Package Option",
        "Selling Price",
        "Ingredient",
        "Quantity",
        "Unit",
        "Cost Per Unit",
        "Ingredient Cost",
    ]
    ws.append(headers)
    style_header_row(ws)

    row_idx = 2
    for variant_idx, variant in enumerate(variants, 2):
        for usage_col_idx, usage_header in enumerate(RECIPE_USAGE_HEADERS[3:], 4):
            ingredient, unit = usage_header.rsplit(" (", 1)
            unit = unit.rstrip(")")
            ws.cell(row=row_idx, column=1, value=variant["product"])
            ws.cell(row=row_idx, column=2, value=variant["package_option"])
            ws.cell(row=row_idx, column=3, value=f"='Recipe Usage'!C{variant_idx}")
            ws.cell(row=row_idx, column=4, value=ingredient)
            ws.cell(
                row=row_idx,
                column=5,
                value=f"='Recipe Usage'!{get_column_letter(usage_col_idx)}{variant_idx}",
            )
            ws.cell(row=row_idx, column=6, value=unit)
            ws.cell(
                row=row_idx,
                column=7,
                value=(
                    f'=IFERROR(VLOOKUP(D{row_idx},'
                    f"'Ingredients'!$A$2:$E${len(INGREDIENTS) + 1},5,FALSE),\"\")"
                ),
            )
            ws.cell(
                row=row_idx,
                column=8,
                value=f'=IF(OR(E{row_idx}=0,G{row_idx}=""),0,E{row_idx}*G{row_idx})',
            )
            for col in [3, 5, 7, 8]:
                ws.cell(row=row_idx, column=col).fill = formula_fill
            row_idx += 1

    last_row = row_idx - 1
    add_table(ws, "CostBreakdownTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [28, 16, 14, 22, 14, 10, 16, 18])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=7).number_format = "#,##0.0000"
        ws.cell(row=row, column=8).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_summary(ws, variants):
    headers = [
        "Product",
        "Package Option",
        "Selling Price",
        "Food Cost",
        "Packaging Cost",
        "Capital Per Serving",
        "Profit",
        "Margin %",
        "Capital % of Price",
        "Missing Cost Inputs",
        "Status",
    ]
    ws.append(headers)
    style_header_row(ws)

    for row_idx, variant in enumerate(variants, 2):
        ws.cell(row=row_idx, column=1, value=variant["product"])
        ws.cell(row=row_idx, column=2, value=variant["package_option"])
        ws.cell(row=row_idx, column=3, value=f"='Recipe Usage'!C{row_idx}")
        ws.cell(
            row=row_idx,
            column=4,
            value=(
                f'=SUMIFS(\'Cost Breakdown\'!$H:$H,\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},\'Cost Breakdown\'!$D:$D,"<>Takeout Box",'
                f'\'Cost Breakdown\'!$D:$D,"<>Dine-in Box")'
            ),
        )
        ws.cell(
            row=row_idx,
            column=5,
            value=(
                f'=SUMIFS(\'Cost Breakdown\'!$H:$H,\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},\'Cost Breakdown\'!$D:$D,"Takeout Box")'
                f'+SUMIFS(\'Cost Breakdown\'!$H:$H,\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},\'Cost Breakdown\'!$D:$D,"Dine-in Box")'
            ),
        )
        ws.cell(row=row_idx, column=6, value=f"=D{row_idx}+E{row_idx}")
        ws.cell(row=row_idx, column=7, value=f'=IF(C{row_idx}="","",C{row_idx}-F{row_idx})')
        ws.cell(
            row=row_idx,
            column=8,
            value=f'=IF(OR(C{row_idx}="",C{row_idx}=0),"",G{row_idx}/C{row_idx})',
        )
        ws.cell(
            row=row_idx,
            column=9,
            value=f'=IF(OR(C{row_idx}="",C{row_idx}=0),"",F{row_idx}/C{row_idx})',
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=(
                f'=COUNTIFS(\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},'
                f'\'Cost Breakdown\'!$E:$E,">0",'
                f'\'Cost Breakdown\'!$G:$G,"")'
            ),
        )
        ws.cell(row=row_idx, column=11, value=f'=IF(J{row_idx}=0,"Complete","Enter missing unit costs")')
        for col in range(3, 12):
            ws.cell(row=row_idx, column=col).fill = formula_fill

    last_row = len(variants) + 1
    add_table(ws, "ProductSummaryTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [28, 16, 14, 14, 16, 20, 14, 12, 18, 18, 20])
    for row in range(2, last_row + 1):
        for col in [3, 4, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = "#,##0.00"
        for col in [8, 9]:
            ws.cell(row=row, column=col).number_format = "0.00%"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_margin_pricing(ws, variants):
    headers = ["Product", "Package Option", "Capital Per Serving"]
    for margin in MARGINS:
        label = f"{int(margin * 100)}% Margin"
        headers.extend([f"{label} Exact Price", f"{label} Rounded Up"])
    ws.append(headers)
    style_header_row(ws)

    for row_idx, variant in enumerate(variants, 2):
        ws.cell(row=row_idx, column=1, value=variant["product"])
        ws.cell(row=row_idx, column=2, value=variant["package_option"])
        ws.cell(row=row_idx, column=3, value=f"='Product Summary'!F{row_idx}")
        ws.cell(row=row_idx, column=3).fill = formula_fill
        col_idx = 4
        for margin in MARGINS:
            exact_cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=f'=IF($C{row_idx}="","",$C{row_idx}/(1-{margin}))',
            )
            rounded_cell = ws.cell(
                row=row_idx,
                column=col_idx + 1,
                value=f'=IF({get_column_letter(col_idx)}{row_idx}="","",ROUNDUP({get_column_letter(col_idx)}{row_idx},0))',
            )
            exact_cell.fill = formula_fill
            rounded_cell.fill = formula_fill
            col_idx += 2

    last_row = len(variants) + 1
    add_table(ws, "MarginPricingTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [28, 16, 20] + [18, 18] * len(MARGINS))
    for row in range(2, last_row + 1):
        for col in range(3, len(headers) + 1):
            ws.cell(row=row, column=col).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    build_workbook()
