from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


OUTPUT_FILE = "general_product_costing_system.xlsx"
MAX_ITEMS = 200
MAX_PRODUCTS = 200
MAX_BOM_LINES = 1000


UNITS = [
    ("g", "Weight", "g", 1),
    ("kg", "Weight", "g", 1000),
    ("mg", "Weight", "g", 0.001),
    ("oz", "Weight", "g", 28.3495),
    ("lb", "Weight", "g", 453.592),
    ("ml", "Volume", "ml", 1),
    ("L", "Volume", "ml", 1000),
    ("fl oz", "Volume", "ml", 29.5735),
    ("pc", "Count", "pc", 1),
    ("pcs", "Count", "pc", 1),
    ("piece", "Count", "pc", 1),
    ("dozen", "Count", "pc", 12),
]


ITEMS = []
PRODUCTS = []
ITEM_IDS = {}
RECIPES = {}


thin_gray = Side(style="thin", color="D9E2F3")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
header_fill = PatternFill("solid", fgColor="1F4E78")
input_fill = PatternFill("solid", fgColor="FFF2CC")
formula_fill = PatternFill("solid", fgColor="E2F0D9")
note_fill = PatternFill("solid", fgColor="FCE4D6")
white_font = Font(color="FFFFFF", bold=True)
title_font = Font(size=16, bold=True, color="1F4E78")
section_font = Font(size=12, bold=True, color="1F4E78")


def add_table(ws, name, start_row, end_row, end_col):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_header_row(ws, max_col):
    for cell in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col).__next__():
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_list_validation(ws, cell_range, formula_name):
    validation = DataValidation(type="list", formula1=f"={formula_name}", allow_blank=True)
    validation.error = "Choose a value from the list or add the value to the related master sheet."
    validation.errorTitle = "Invalid value"
    ws.add_data_validation(validation)
    validation.add(cell_range)


def set_common_sheet_options(ws, freeze_cell="A2"):
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    ws_intro = wb.create_sheet("Instructions")
    ws_units = wb.create_sheet("Units")
    ws_items = wb.create_sheet("Items")
    ws_products = wb.create_sheet("Products")
    ws_bom = wb.create_sheet("Recipe BOM")
    ws_summary = wb.create_sheet("Product Summary")

    build_instructions(ws_intro)
    build_units(ws_units)
    build_items(ws_items)
    build_products(ws_products)
    build_recipe_bom(ws_bom)
    build_summary(ws_summary)
    add_defined_names_and_validations(wb, ws_items, ws_products, ws_bom)

    wb.active = wb.sheetnames.index("Instructions")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT_FILE)
    load_workbook(OUTPUT_FILE, data_only=False)


def build_instructions(ws):
    ws["A1"] = "General Product Costing System"
    ws["A1"].font = title_font

    ws["A3"] = "How the system works"
    ws["A3"].font = section_font
    lines = [
        "1. Update Units if you need more units or conversion factors.",
        "2. Add or update ingredients, packaging, materials, or supplies in Items.",
        "3. Add products and selling prices in Products.",
        "4. Add each product's ingredients/materials in Recipe BOM.",
        "5. Product Summary automatically calculates capital, gross profit, margin, and issues.",
    ]
    for row_idx, text in enumerate(lines, 4):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A11"] = "Important"
    ws["A11"].font = section_font
    notes = [
        "Item costs are entered once in Items. Any recipe using that item updates when the item purchase cost changes.",
        "Product selling prices are entered once in Products. Product Summary updates when the price changes.",
        "This workbook starts empty, so you can build a new list of items, products, and recipes from scratch.",
        "Recipe units can differ from purchase units if both convert to the item's base unit. Example: buy flour by kg, use flour by g.",
        "Use the same currency everywhere.",
        "Rows are prebuilt for 200 items, 200 products, and 1000 recipe/BOM lines. Copy formulas down if you need more.",
    ]
    for row_idx, text in enumerate(notes, 12):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A19"] = "Color guide"
    ws["A19"].font = section_font
    guide = [
        ("Yellow", "Input cells you edit."),
        ("Green", "Formula cells that calculate automatically."),
        ("Orange", "Notes and guidance."),
        ("Blue", "Table headers."),
    ]
    for row_idx, (label, text) in enumerate(guide, 20):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=text)
    ws["A20"].fill = input_fill
    ws["A21"].fill = formula_fill
    ws["A22"].fill = note_fill
    ws["A23"].fill = header_fill
    ws["A23"].font = white_font

    set_column_widths(ws, [24, 96])
    style_range(ws, 1, 23, 1, 2)
    ws.sheet_view.showGridLines = False


def build_units(ws):
    headers = ["Unit", "Unit Type", "Base Unit", "Factor to Base Unit", "Notes"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    for unit, unit_type, base_unit, factor in UNITS:
        ws.append([unit, unit_type, base_unit, factor, f"1 {unit} = {factor} {base_unit}"])
    last_row = len(UNITS) + 1
    add_table(ws, "UnitsTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [14, 14, 14, 20, 32])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=4).number_format = "#,##0.0000"
    set_common_sheet_options(ws)


def build_items(ws):
    headers = [
        "Item ID",
        "Item Name",
        "Category",
        "Base Unit",
        "Purchase Qty",
        "Purchase Unit",
        "Purchase Cost",
        "Purchase Qty in Base Unit",
        "Waste %",
        "Cost Per Base Unit",
        "Status",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    units_ref = f"'Units'!$A$2:$D${len(UNITS) + 1}"
    for row_idx in range(2, MAX_ITEMS + 2):
        if row_idx - 2 < len(ITEMS):
            item_id, name, category, base_unit, purchase_qty, purchase_unit, purchase_cost, waste, notes = ITEMS[row_idx - 2]
            base_values = [item_id, name, category, base_unit, purchase_qty, purchase_unit, purchase_cost]
            for col_idx, value in enumerate(base_values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            ws.cell(row=row_idx, column=9, value=waste)
            ws.cell(row=row_idx, column=12, value=notes)
        ws.cell(row=row_idx, column=8, value=f'=IF(OR(E{row_idx}="",F{row_idx}=""),"",IFERROR(E{row_idx}*VLOOKUP(F{row_idx},{units_ref},4,FALSE),""))')
        ws.cell(row=row_idx, column=10, value=f'=IF(OR(G{row_idx}="",H{row_idx}=""),"",IFERROR(G{row_idx}/(H{row_idx}*(1-I{row_idx})),""))')
        ws.cell(
            row=row_idx,
            column=11,
            value=(
                f'=IF(AND(A{row_idx}="",B{row_idx}=""),"",'
                f'IF(OR(A{row_idx}="",B{row_idx}="",D{row_idx}="",E{row_idx}="",F{row_idx}="",G{row_idx}=""),"Missing item data",'
                f'IF(IFERROR(VLOOKUP(F{row_idx},{units_ref},3,FALSE),"")<>D{row_idx},"Unit mismatch",'
                f'IF(J{row_idx}="","Check cost","OK"))))'
            ),
        )

    last_row = MAX_ITEMS + 1
    add_table(ws, "ItemsTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [16, 26, 18, 12, 14, 16, 14, 24, 12, 20, 18, 46])
    for row in range(2, last_row + 1):
        for col in [5, 7, 8, 10]:
            ws.cell(row=row, column=col).number_format = "#,##0.0000"
        ws.cell(row=row, column=9).number_format = "0.00%"
        for col in [4, 5, 6, 7, 9]:
            ws.cell(row=row, column=col).fill = input_fill
        for col in [8, 10, 11]:
            ws.cell(row=row, column=col).fill = formula_fill
        ws.cell(row=row, column=12).fill = note_fill
    set_common_sheet_options(ws)


def build_products(ws):
    headers = ["Product ID", "Product Name", "Category", "Selling Price", "Active?", "Notes"]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for row_idx in range(2, MAX_PRODUCTS + 2):
        if row_idx - 2 < len(PRODUCTS):
            product_id, name, category, price = PRODUCTS[row_idx - 2]
            values = [product_id, name, category, price, "Yes", ""]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col in [1, 2, 3, 4, 5, 6]:
            ws.cell(row=row_idx, column=col).fill = input_fill

    last_row = MAX_PRODUCTS + 1
    add_table(ws, "ProductsTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [26, 28, 18, 16, 12, 44])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=4).number_format = "#,##0.00"
    set_common_sheet_options(ws)


def build_recipe_bom(ws):
    headers = [
        "Line ID",
        "Product ID",
        "Product Name",
        "Item ID",
        "Item Name",
        "Quantity",
        "Recipe Unit",
        "Qty in Item Base Unit",
        "Item Base Unit",
        "Cost Per Base Unit",
        "Line Cost",
        "Status",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    item_ref = f"'Items'!$A$2:$L${MAX_ITEMS + 1}"
    product_ref = f"'Products'!$A$2:$F${MAX_PRODUCTS + 1}"
    units_ref = f"'Units'!$A$2:$D${len(UNITS) + 1}"

    initial_lines = []
    for product_id, recipe_lines in RECIPES.items():
        for item_name, quantity, unit in recipe_lines:
            initial_lines.append((product_id, ITEM_IDS[item_name], quantity, unit, ""))

    for row_idx in range(2, MAX_BOM_LINES + 2):
        ws.cell(row=row_idx, column=1, value=f'=IF(AND(B{row_idx}="",D{row_idx}=""),"",ROW()-1)')
        if row_idx - 2 < len(initial_lines):
            product_id, item_id, quantity, unit, note = initial_lines[row_idx - 2]
            ws.cell(row=row_idx, column=2, value=product_id)
            ws.cell(row=row_idx, column=4, value=item_id)
            ws.cell(row=row_idx, column=6, value=quantity)
            ws.cell(row=row_idx, column=7, value=unit)
            ws.cell(row=row_idx, column=13, value=note)
        ws.cell(row=row_idx, column=3, value=f'=IF(B{row_idx}="","",IFERROR(VLOOKUP(B{row_idx},{product_ref},2,FALSE),"Product not found"))')
        ws.cell(row=row_idx, column=5, value=f'=IF(D{row_idx}="","",IFERROR(VLOOKUP(D{row_idx},{item_ref},2,FALSE),"Item not found"))')
        ws.cell(row=row_idx, column=8, value=f'=IF(OR(F{row_idx}="",G{row_idx}=""),"",IFERROR(F{row_idx}*VLOOKUP(G{row_idx},{units_ref},4,FALSE),""))')
        ws.cell(row=row_idx, column=9, value=f'=IF(D{row_idx}="","",IFERROR(VLOOKUP(D{row_idx},{item_ref},4,FALSE),""))')
        ws.cell(row=row_idx, column=10, value=f'=IF(D{row_idx}="","",IFERROR(VLOOKUP(D{row_idx},{item_ref},10,FALSE),""))')
        ws.cell(row=row_idx, column=11, value=f'=IF(OR(H{row_idx}="",J{row_idx}=""),"",H{row_idx}*J{row_idx})')
        ws.cell(
            row=row_idx,
            column=12,
            value=(
                f'=IF(AND(B{row_idx}="",D{row_idx}=""),"",'
                f'IF(B{row_idx}="","Missing product",'
                f'IF(D{row_idx}="","Missing item",'
                f'IF(C{row_idx}="Product not found","Product not found",'
                f'IF(E{row_idx}="Item not found","Item not found",'
                f'IF(OR(F{row_idx}="",G{row_idx}=""),"Missing quantity/unit",'
                f'IF(IFERROR(VLOOKUP(G{row_idx},{units_ref},3,FALSE),"")<>I{row_idx},"Unit mismatch",'
                f'IF(J{row_idx}="","Enter item cost","OK"))))))))'
            ),
        )

    last_row = MAX_BOM_LINES + 1
    add_table(ws, "RecipeBOMTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [10, 26, 28, 16, 26, 12, 14, 22, 14, 18, 16, 20, 36])
    for row in range(2, last_row + 1):
        for col in [6, 8, 10, 11]:
            ws.cell(row=row, column=col).number_format = "#,##0.0000"
        for col in [2, 4, 6, 7, 13]:
            ws.cell(row=row, column=col).fill = input_fill
        for col in [1, 3, 5, 8, 9, 10, 11, 12]:
            ws.cell(row=row, column=col).fill = formula_fill
    set_common_sheet_options(ws)


def build_summary(ws):
    headers = [
        "Product ID",
        "Product Name",
        "Category",
        "Selling Price",
        "Capital / Total Cost",
        "Gross Profit",
        "Profit Margin",
        "Cost % of Price",
        "Recipe Lines",
        "Issues",
        "Status",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for row_idx in range(2, MAX_PRODUCTS + 2):
        product_row = row_idx
        ws.cell(row=row_idx, column=1, value=f'=IF(Products!A{product_row}="","",Products!A{product_row})')
        ws.cell(row=row_idx, column=2, value=f'=IF(A{row_idx}="","",Products!B{product_row})')
        ws.cell(row=row_idx, column=3, value=f'=IF(A{row_idx}="","",Products!C{product_row})')
        ws.cell(row=row_idx, column=4, value=f'=IF(A{row_idx}="","",Products!D{product_row})')
        ws.cell(row=row_idx, column=5, value=f'=IF(A{row_idx}="","",SUMIFS(\'Recipe BOM\'!$K:$K,\'Recipe BOM\'!$B:$B,A{row_idx}))')
        ws.cell(row=row_idx, column=6, value=f'=IF(OR(A{row_idx}="",D{row_idx}="",E{row_idx}=""),"",D{row_idx}-E{row_idx})')
        ws.cell(row=row_idx, column=7, value=f'=IF(OR(A{row_idx}="",D{row_idx}="",E{row_idx}=""),"",F{row_idx}/D{row_idx})')
        ws.cell(row=row_idx, column=8, value=f'=IF(OR(A{row_idx}="",D{row_idx}="",E{row_idx}=""),"",E{row_idx}/D{row_idx})')
        ws.cell(row=row_idx, column=9, value=f'=IF(A{row_idx}="","",COUNTIFS(\'Recipe BOM\'!$B:$B,A{row_idx},\'Recipe BOM\'!$D:$D,"<>"))')
        ws.cell(
            row=row_idx,
            column=10,
            value=f'=IF(A{row_idx}="","",COUNTIFS(\'Recipe BOM\'!$B:$B,A{row_idx},\'Recipe BOM\'!$D:$D,"<>",\'Recipe BOM\'!$L:$L,"<>OK"))',
        )
        ws.cell(
            row=row_idx,
            column=11,
            value=(
                f'=IF(A{row_idx}="","",'
                f'IF(I{row_idx}=0,"No recipe",'
                f'IF(J{row_idx}>0,"Check recipe/items",'
                f'IF(D{row_idx}="","Enter selling price","OK"))))'
            ),
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = formula_fill

    last_row = MAX_PRODUCTS + 1
    add_table(ws, "ProductSummaryTable", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [26, 28, 18, 16, 20, 16, 16, 16, 14, 12, 20])
    for row in range(2, last_row + 1):
        for col in [4, 5, 6]:
            ws.cell(row=row, column=col).number_format = "#,##0.00"
        for col in [7, 8]:
            ws.cell(row=row, column=col).number_format = "0.00%"
    set_common_sheet_options(ws)


def add_defined_names_and_validations(wb, ws_items, ws_products, ws_bom):
    defined_names = {
        "UnitList": f"'Units'!$A$2:$A${len(UNITS) + 1}",
        "ItemIDList": f"'Items'!$A$2:$A${MAX_ITEMS + 1}",
        "ProductIDList": f"'Products'!$A$2:$A${MAX_PRODUCTS + 1}",
    }
    for name, target in defined_names.items():
        wb.defined_names.add(DefinedName(name, attr_text=target))

    add_list_validation(ws_items, f"D2:D{MAX_ITEMS + 1}", "UnitList")
    add_list_validation(ws_items, f"F2:F{MAX_ITEMS + 1}", "UnitList")
    add_list_validation(ws_bom, f"B2:B{MAX_BOM_LINES + 1}", "ProductIDList")
    add_list_validation(ws_bom, f"D2:D{MAX_BOM_LINES + 1}", "ItemIDList")
    add_list_validation(ws_bom, f"G2:G{MAX_BOM_LINES + 1}", "UnitList")


if __name__ == "__main__":
    build_workbook()
