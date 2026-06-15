from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_FILE = "coffee_capital_system.xlsx"


INGREDIENTS = [
    ("Coffee", "g", "Enter coffee bag total grams and purchase cost."),
    ("Milk", "ml", "Enter carton/bottle total ml and purchase cost."),
    ("Cup", "pc", "Enter number of iced/cold cups per pack and pack cost."),
    ("Lid", "pc", "Enter number of iced/cold lids per pack and pack cost."),
    ("Hot Cup", "pc", "Enter number of hot cups per pack and pack cost."),
    ("Hot Lid", "pc", "Enter number of hot lids per pack and pack cost."),
    ("Ice", "g", "Enter total grams bought/produced and cost."),
    ("Straw", "pc", "Enter number of straws per pack and pack cost."),
    ("Water", "ml", "Enter total ml and cost if you want to include water."),
    ("Condensed milk", "g", "Enter can/container total grams and cost."),
    ("Ube powder", "g", "Enter pouch/container total grams and cost."),
    ("Caramel Syrup", "g", "Enter bottle total grams and cost."),
    ("Chocolate Syrup", "g", "Enter bottle total grams and cost."),
]


RECIPE_USAGE_HEADERS = [
    "Type",
    "Recipe",
    "Selling Price",
    "Coffee (g)",
    "Milk (ml)",
    "Water (ml)",
    "Ice (g)",
    "Condensed milk (g)",
    "Ube powder (g)",
    "Caramel Syrup (g)",
    "Chocolate Syrup (g)",
    "Cup (pc)",
    "Lid (pc)",
    "Hot Cup (pc)",
    "Hot Lid (pc)",
    "Straw (pc)",
]


RECIPES = [
    {
        "type": "Iced",
        "recipe": "Spanish latte",
        "price": 105,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 150,
            "Ice (g)": 100,
            "Condensed milk (g)": 30,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Americano",
        "price": 70,
        "usage": {
            "Coffee (g)": 18,
            "Water (ml)": 150,
            "Ice (g)": 100,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Latte",
        "price": 90,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 150,
            "Ice (g)": 100,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Ube Latte",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 150,
            "Ice (g)": 100,
            "Condensed milk (g)": 30,
            "Ube powder (g)": 8,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Mocha Latte",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 150,
            "Ice (g)": 100,
            "Condensed milk (g)": 30,
            "Chocolate Syrup (g)": 35,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Caramel Macchiato",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 150,
            "Ice (g)": 100,
            "Condensed milk (g)": 30,
            "Caramel Syrup (g)": 35,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Iced",
        "recipe": "Kape Sua Da",
        "price": 105,
        "usage": {
            "Coffee (g)": 18,
            "Water (ml)": 150,
            "Ice (g)": 100,
            "Condensed milk (g)": 30,
            "Cup (pc)": 1,
            "Lid (pc)": 1,
            "Straw (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Spanish latte",
        "price": 105,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 200,
            "Condensed milk (g)": 30,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Americano",
        "price": 70,
        "usage": {
            "Coffee (g)": 18,
            "Water (ml)": 200,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Latte",
        "price": 90,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 200,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Ube Latte",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 200,
            "Condensed milk (g)": 30,
            "Ube powder (g)": 8,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Mocha Latte",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 200,
            "Condensed milk (g)": 30,
            "Chocolate Syrup (g)": 35,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
    {
        "type": "Hot",
        "recipe": "Caramel Macchiato",
        "price": 110,
        "usage": {
            "Coffee (g)": 18,
            "Milk (ml)": 200,
            "Condensed milk (g)": 30,
            "Caramel Syrup (g)": 35,
            "Hot Cup (pc)": 1,
            "Hot Lid (pc)": 1,
        },
    },
]


BREAKDOWN_HEADERS = [
    "Type",
    "Recipe",
    "Selling Price",
    "Ingredient",
    "Quantity",
    "Unit",
    "Cost Per Unit",
    "Ingredient Cost",
]


SUMMARY_HEADERS = [
    "Type",
    "Recipe",
    "Selling Price",
    "Capital Per Cup",
    "Gross Profit Per Cup",
    "Profit Margin",
    "Capital % of Price",
    "Missing Cost Inputs",
    "Status",
]


thin_gray = Side(style="thin", color="D9E2F3")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
header_fill = PatternFill("solid", fgColor="1F4E78")
subheader_fill = PatternFill("solid", fgColor="5B9BD5")
input_fill = PatternFill("solid", fgColor="FFF2CC")
formula_fill = PatternFill("solid", fgColor="E2F0D9")
note_fill = PatternFill("solid", fgColor="FCE4D6")
white_font = Font(color="FFFFFF", bold=True)
title_font = Font(size=16, bold=True, color="1F4E78")
section_font = Font(size=12, bold=True, color="1F4E78")


def style_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(ws, name, start_row, end_row, end_col):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_workbook():
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_intro = wb.create_sheet("Instructions")
    ws_costs = wb.create_sheet("Unit Costs")
    ws_usage = wb.create_sheet("Recipe Usage")
    ws_breakdown = wb.create_sheet("Cost Breakdown")
    ws_summary = wb.create_sheet("Recipe Summary")

    build_instructions(ws_intro)
    build_unit_costs(ws_costs)
    build_recipe_usage(ws_usage)
    build_cost_breakdown(ws_breakdown)
    build_summary(ws_summary)

    wb.active = wb.sheetnames.index("Instructions")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT_FILE)

    # Validate that the workbook can be loaded after writing.
    load_workbook(OUTPUT_FILE, data_only=False)


def build_instructions(ws):
    ws["A1"] = "Coffee Capital System"
    ws["A1"].font = title_font
    ws["A3"] = "How to use"
    ws["A3"].font = section_font
    instructions = [
        "1. Go to the Unit Costs sheet.",
        "2. Enter each ingredient's Purchase Qty and Purchase Cost in the yellow cells.",
        "3. Review Recipe Summary for Capital Per Cup, Gross Profit, Profit Margin, and Capital % of Price.",
        "4. If you change a recipe quantity, update the Recipe Usage sheet. The Cost Breakdown and Recipe Summary formulas will recalculate.",
        "5. Use the Missing Cost Inputs and Status columns to see whether a recipe has all required cost data.",
    ]
    for row_idx, text in enumerate(instructions, 4):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A11"] = "Important assumptions"
    ws["A11"].font = section_font
    assumptions = [
        "Selling prices are entered exactly as provided.",
        "Quantities are counted only when the ingredient appears in the provided recipe list.",
        "Iced recipes use the Cup and Lid cost inputs; hot recipes use the separate Hot Cup and Hot Lid cost inputs.",
        "Hot coffee base data mentioned Straw, but hot recipes did not list Straw; hot recipe straw quantity is set to 0. Change Recipe Usage if you include a straw or stirrer.",
        "Water cost can be entered as 0 if you do not track it separately.",
        "All costs should be entered in the same currency as your selling prices.",
    ]
    for row_idx, text in enumerate(assumptions, 12):
        ws.cell(row=row_idx, column=1, value=text)

    ws["A19"] = "Formula color guide"
    ws["A19"].font = section_font
    legend = [
        ("Yellow", "Input cells: enter purchase quantity and cost."),
        ("Green", "Formula cells: calculated automatically."),
        ("Blue", "Table headers."),
        ("Orange", "Notes and assumptions."),
    ]
    for row_idx, (label, text) in enumerate(legend, 20):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=text)
    ws["A20"].fill = input_fill
    ws["A21"].fill = formula_fill
    ws["A22"].fill = header_fill
    ws["A22"].font = white_font
    ws["A23"].fill = note_fill

    set_column_widths(ws, [24, 90])
    style_range(ws, 1, 23, 1, 2)
    ws.sheet_view.showGridLines = False


def build_unit_costs(ws):
    headers = [
        "Ingredient",
        "Measurement Unit",
        "Purchase Qty",
        "Purchase Cost",
        "Cost Per Unit",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (ingredient, unit, note) in enumerate(INGREDIENTS, 2):
        ws.cell(row=row_idx, column=1, value=ingredient)
        ws.cell(row=row_idx, column=2, value=unit)
        ws.cell(row=row_idx, column=3).fill = input_fill
        ws.cell(row=row_idx, column=4).fill = input_fill
        ws.cell(row=row_idx, column=5, value=f'=IF(OR(C{row_idx}="",D{row_idx}=""),"",D{row_idx}/C{row_idx})')
        ws.cell(row=row_idx, column=5).fill = formula_fill
        ws.cell(row=row_idx, column=6, value=note)
        ws.cell(row=row_idx, column=6).fill = note_fill

    last_row = len(INGREDIENTS) + 1
    add_table(ws, "UnitCosts", 1, last_row, len(headers))
    style_range(ws, 1, last_row, 1, len(headers))
    set_column_widths(ws, [22, 18, 16, 16, 16, 54])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.0000"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_recipe_usage(ws):
    ws.append(RECIPE_USAGE_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for recipe in RECIPES:
        row = [recipe["type"], recipe["recipe"], recipe["price"]]
        for header in RECIPE_USAGE_HEADERS[3:]:
            row.append(recipe["usage"].get(header, 0))
        ws.append(row)

    last_row = len(RECIPES) + 1
    add_table(ws, "RecipeUsage", 1, last_row, len(RECIPE_USAGE_HEADERS))
    style_range(ws, 1, last_row, 1, len(RECIPE_USAGE_HEADERS))
    set_column_widths(
        ws,
        [12, 22, 14, 12, 12, 12, 10, 20, 16, 18, 20, 10, 10, 12, 12, 10],
    )
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        for col in range(4, len(RECIPE_USAGE_HEADERS) + 1):
            ws.cell(row=row, column=col).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_cost_breakdown(ws):
    ws.append(BREAKDOWN_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for recipe in RECIPES:
        for ingredient_header in RECIPE_USAGE_HEADERS[3:]:
            quantity = recipe["usage"].get(ingredient_header, 0)
            if quantity == 0:
                continue
            ingredient, unit = ingredient_header.rsplit(" (", 1)
            unit = unit.rstrip(")")
            ws.cell(row=row_idx, column=1, value=recipe["type"])
            ws.cell(row=row_idx, column=2, value=recipe["recipe"])
            ws.cell(row=row_idx, column=3, value=recipe["price"])
            ws.cell(row=row_idx, column=4, value=ingredient)
            ws.cell(row=row_idx, column=5, value=quantity)
            ws.cell(row=row_idx, column=6, value=unit)
            ws.cell(
                row=row_idx,
                column=7,
                value=f'=IFERROR(VLOOKUP(D{row_idx},\'Unit Costs\'!$A$2:$E${len(INGREDIENTS) + 1},5,FALSE),"")',
            )
            ws.cell(
                row=row_idx,
                column=8,
                value=f'=IF(G{row_idx}="","",E{row_idx}*G{row_idx})',
            )
            ws.cell(row=row_idx, column=7).fill = formula_fill
            ws.cell(row=row_idx, column=8).fill = formula_fill
            row_idx += 1

    last_row = row_idx - 1
    add_table(ws, "CostBreakdown", 1, last_row, len(BREAKDOWN_HEADERS))
    style_range(ws, 1, last_row, 1, len(BREAKDOWN_HEADERS))
    set_column_widths(ws, [12, 22, 14, 22, 12, 10, 16, 18])
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=7).number_format = "#,##0.0000"
        ws.cell(row=row, column=8).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_summary(ws):
    ws.append(SUMMARY_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, recipe in enumerate(RECIPES, 2):
        ws.cell(row=row_idx, column=1, value=recipe["type"])
        ws.cell(row=row_idx, column=2, value=recipe["recipe"])
        ws.cell(row=row_idx, column=3, value=recipe["price"])
        ws.cell(
            row=row_idx,
            column=4,
            value=(
                f'=IF(COUNTIFS(\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},\'Cost Breakdown\'!$G:$G,"")>0,"",'
                f'SUMIFS(\'Cost Breakdown\'!$H:$H,\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx}))'
            ),
        )
        ws.cell(row=row_idx, column=5, value=f'=IF(D{row_idx}="","",C{row_idx}-D{row_idx})')
        ws.cell(row=row_idx, column=6, value=f'=IF(D{row_idx}="","",E{row_idx}/C{row_idx})')
        ws.cell(row=row_idx, column=7, value=f'=IF(D{row_idx}="","",D{row_idx}/C{row_idx})')
        ws.cell(
            row=row_idx,
            column=8,
            value=(
                f'=COUNTIFS(\'Cost Breakdown\'!$A:$A,A{row_idx},'
                f'\'Cost Breakdown\'!$B:$B,B{row_idx},\'Cost Breakdown\'!$G:$G,"")'
            ),
        )
        ws.cell(row=row_idx, column=9, value=f'=IF(H{row_idx}=0,"Complete","Enter unit costs")')
        for col in range(4, 10):
            ws.cell(row=row_idx, column=col).fill = formula_fill

    last_row = len(RECIPES) + 1
    add_table(ws, "RecipeSummary", 1, last_row, len(SUMMARY_HEADERS))
    style_range(ws, 1, last_row, 1, len(SUMMARY_HEADERS))
    set_column_widths(ws, [12, 22, 14, 18, 20, 16, 18, 18, 18])
    for row in range(2, last_row + 1):
        for col in [3, 4, 5]:
            ws.cell(row=row, column=col).number_format = "#,##0.00"
        for col in [6, 7]:
            ws.cell(row=row, column=col).number_format = "0.00%"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    build_workbook()
